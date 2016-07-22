import fnmatch
import codecs
import tablib
import os
import requests
import urlparse
import shutil
import frontmatter
import yaml
import io

class PrizeSpreadsheets(object):
    """
    """

    def __init__(self, dir_path):
        self.dir_path = dir_path

    def get_spreadsheets(self):
        sheets = []
        for file in os.listdir(self.dir_path):
            if file.endswith(".xlsx") and file.startswith("~") == False:
                new_filename = "%s.xlsx" % (self.get_region(file))
                os.rename(os.path.join(self.dir_path, file), os.path.join(self.dir_path, new_filename))
                # print new_filename

                sheets.append({
                    "filename": new_filename,
                    "region": self.get_region(file),
                    "sheet": PrizeSpreadsheet(os.path.join(self.dir_path, new_filename))
                })
        return sheets
    
    def get_region(self, file):
        regions = ["NATIONAL", "ACT", "NSW", "NT", "QLD", "SA", "TAS", "VIC", "WA"]

        if " " not in file:
            region = os.path.splitext(file)[0]
        else:
            region = file.split(" ")[0].upper()

        if region not in regions:
            raise ValueError("Region '%s' is not valid." % (region))
        
        if region == "NATIONAL":
            return "AUSTRALIA"

        return region

class PrizeSpreadsheet(object):
    """
    """
    
    def __init__(self, file_path):
        self.file_path = file_path
        self.read_file()
    
    def read_file(self):
        from openpyxl import load_workbook
        # print "file_path: %s" % (self.file_path)
        self.wb = load_workbook(self.file_path)
    
    def get_prize_sheet(self, index=0):
        # print self.wb.get_sheet_names()
        self.ws = self.wb.get_sheet_by_name(self.wb.get_sheet_names()[index])
        return self.ws
    
    def get_headers(self):
        headers = []
        for row in self.ws.iter_rows("A1:Z1"):
            for cell in row:
                if cell.value == None:
                    break
                else:
                    col_name = "".join(cell.value.split()).lower()
                    if "eligibilitycriteria" in col_name or "eligiblitycriteria" in col_name:
                        col_name = "eligibilitycriteria"
                    headers.append(col_name)
        return headers

    def is_row_empty(self, row):
        for cell in row:
            if cell.value is not None:
                return False
        return True

    def sheet_to_dict(self, ws):
        rows = []
        headers = self.get_headers()
        print headers
        # print
        
        for row in ws.iter_rows(row_offset=1):
            # print row

            # Consider a single empty row to be the end of the available data
            if self.is_row_empty(row) == True:
                # print "Row is empty"
                break

            tmp = {}
            for idx, cell in enumerate(row):
                # print "%s: %s" % (idx, cell.value)
                if idx >= len(headers):
                    break
                else:
                    if type(cell.value) is unicode:
                        tmp[headers[idx]] = cell.value.strip()
                    else:
                        tmp[headers[idx]] = cell.value
            rows.append(tmp)
            # print
        return rows


# Config
datadir = "python/data/prizes"
prizesdir = "_prizes/2016/"
organisationsdir = "_organisations/"
eventsdir = "_locations/"
# tmpdir = "python/tmp/"

# Init
# For seeing if the mentor's organisations exists
organisation_names = []
for root, dirnames, filenames in os.walk(organisationsdir):
    for filename in fnmatch.filter(filenames, "*.md"):
        organisation_names.append(filename.split(".")[0])

# For matching mentors to events
event_names = []
event_md_files = {}
for root, dirnames, filenames in os.walk(eventsdir):
    for filename in fnmatch.filter(filenames, "*.md"):
        event_name = filename.split(".")[0]
        event_names.append(event_name)
        event_md_files[event_name] = os.path.join(root, filename)

# Ingest available prize sheets for UPSERTing
gids = [] # GIDs used so far for sanity checking
validation_errors = []

sheets = PrizeSpreadsheets(datadir).get_spreadsheets()
for file in sheets:
    print "Spreadsheet: %s" % (file["filename"])
    print "Region: %s" % (file["region"])

    ws = file["sheet"].get_prize_sheet(0)
    rows = file["sheet"].sheet_to_dict(ws)

    print "Prize Count: %s" % (len(rows))
    print
    print "Processing prizes..."
    print

    for idx, row in enumerate(rows):
        # Assign our prize a globally unique id
        # if "prizename" not in row:
        #     print row
        #     exit()
        if row["prizename"] is None:
            # @TODO
            errmsg = "%s: Prize #'%s' has no name" % (file["region"].lower(), idx)
            # raise ValueError(errmsg)
            validation_errors.append(errmsg)
            print errmsg
            continue
        
        gid = file["region"].lower() + "-" + row["prizename"].lower().strip().replace("/", " or ").replace(" ", "-").replace("'", "")
        if gid in gids: # Hacky
            gid = gid + "-2"

        if gid in gids:
            raise ValueError("GID '%s' is already in use." % (gid))
        else:
            gids.append(gid)
        
        print row["prizename"]
        print gid
        print

        if row["prizetype"] is None:
            errmsg = "%s: Prize '%s' has no type set." % (file["region"].lower(), row["prizename"])
            # raise ValueError(errmsg)
            validation_errors.append(errmsg)
            print errmsg
            continue

        prize = {
            "name": row["prizename"],
            "title": row["prizename"],
            "gid": gid,
            "jurisdiction": file["region"].lower(),
            "type": row["prizetype"]
            # "type": row["What type of mentor are you?"].strip(),
            # "position_title": row["Job title"].strip(),
            # "ask_me_about": row["What can people ask you about? (in a sentence)"].strip(),
            # "organisation": row["Agency/Organisation"].strip(),
            # "jurisdiction": row["What state or territory do you reside in?"].lower(),
            # "contact": {
            #     "email": row["Email"]
            # }
        }

        # Attach non-national prizes to their events
        if file["region"] != "AUSTRALIA":
            if row["prizelevelregionwideoreventspecific"] == "Event only":
                if "eventspecificlocation" not in row:
                    raise ValueError("Event-only prize nominated without any accompanying event specified.")

                if row["eventspecificlocation"] is None:
                    # @TODO
                    errmsg = "%s: Prize '%s' is an Event prize, but no event locations provided." % (file["region"].lower(), row["prizename"])
                    # raise ValueError(errmsg)
                    validation_errors.append(errmsg)
                    print errmsg
                    continue

                event_gid = row["eventspecificlocation"].replace(" ", "-").replace(",", "").lower()
                event_gid_original = event_gid

                # Hacky fix
                if event_gid == "mount-gambier":
                    event_gid = "mount-gambier-youth"
                elif event_gid == "all-brisbane-events":
                    event_gid = "brisbane"

                if event_gid not in event_names:
                    errmsg = "%s: Event GID '%s' does not exist." % (file["region"].lower(), event_gid)
                    raise ValueError(errmsg)
                    validation_errors.append(errmsg)
                    # print errmsg
                else:
                    print "For Event: %s" % (event_gid)

                    post = frontmatter.load(event_md_files[event_gid])
                    prize["category"] = "local"
                    prize["events"] = [post.metadata["gid"]]

                    # Hacky fix
                    if event_gid_original == "all-brisbane-events":
                        prize["events"].append("brisbane-youth")
                        prize["events"].append("brisbane-maker")
                    
                    if post.metadata["gid"] != event_gid:
                        print "WARNING: Event .md file does not match event gid. %s, %s" % (event_gid, post.metadata["gid"])
            else:
                prize["category"] = "state"
        else:
            prize["category"] = "australia"
            
        # Attach sponsoring organisations
        organisation_gid = row["sponsoredby"].lower().replace(" ", "-").replace(",", "").strip()
        if organisation_gid in organisation_names:
            prize["organisation"] = organisation_gid
        else:
            print "WARNING: Could not resolve organisation: %s (%s)" % (row["sponsoredby"], organisation_gid)

        # If a prize already exists, merge the latest info over the top
        prize_md_dir = os.path.join(prizesdir, file["region"].lower())
        prize_md_file = os.path.join(prize_md_dir, "%s.md" % (gid))

        if os.path.exists(prize_md_file):
            # print "NOTICE: Found an existing prize. Merging new data."
            existing_prize = frontmatter.load(prize_md_file)
            existing_prize.metadata.update(prize)
            prize = existing_prize.metadata

        # Convert prize $$$ value to an integer
        estimatedprizevalue = ""
        if type(row["estimateprizevalue$"]) is unicode and row["estimateprizevalue$"].strip().replace("$", "").isdigit():
            estimatedprizevalue = int(row["estimateprizevalue$"].strip().replace("$", ""))
        elif type(row["estimateprizevalue$"]) is float or type(row["estimateprizevalue$"]) is long:
            estimatedprizevalue = int(row["estimateprizevalue$"])
        elif row["estimateprizevalue$"] is not None:
            estimatedprizevalue = row["estimateprizevalue$"].strip()
        
        # Fixing up minor stuff
        if row["prizecategorydescription"] is None:
            row["prizecategorydescription"] = unicode("")
        if row["prizereward"] is None:
            row["prizereward"] = unicode("")
        if row["eligibilitycriteria"] is None:
            row["eligibilitycriteria"] = unicode("")
        
        # print prize
        # print row
        print
        print "---"
        print

        # continue
        if not os.path.exists(prize_md_dir):
            os.makedirs(prize_md_dir)

        with io.open(prize_md_file, "w", encoding="utf-8") as f:
            f.write(u'---\n')
            f.write(unicode(yaml.safe_dump(prize, width=200, default_flow_style=False, encoding="utf-8", allow_unicode=True), "utf-8"))
            f.write(u'---\n')
            f.write(u'\n')
            f.write(unicode(row["prizecategorydescription"].replace("|", "\n").rstrip()))
            f.write(u'\n\n')
            f.write(u'# Prize\n')
            f.write(unicode(row["prizereward"].replace("|", "\n").rstrip()))
            f.write(u'\n\n')
            f.write(u'## Estimated Prize Value\n')
            f.write(unicode("$%s" % (estimatedprizevalue)))
            f.write(u'\n\n')
            f.write(u'# Eligibility Criteria\n')
            f.write(unicode(row["eligibilitycriteria"].replace("|", "\n").rstrip()))
    
    # print "\n"
    print "############################################################"
    # print "\n\n"

if len(validation_errors) > 0:
    for i in validation_errors:
        print i
# ---
# name: Prize 1
# id: prize_1
# photo_url: https://static.pexels.com/photos/3084/person-woman-park-music-large.jpg
# jurisdiction: australia
# type: Prize
# organisations:
#   - organisation_1
# themes:
#   - theme_1
#   - theme_3
# datasets:
#   - dataset_1
#   - dataset_3
# dataportals:
#   - dataportal_1
# ---

# Lorem ipsum dolor sit amet, consectetur adipiscing elit. Aliquam at ornare risus, at dignissim sapien. Sed eget est mi. Ut lacinia ornare tellus commodo sagittis. Integer euismod eleifend velit, eget dictum leo sagittis at.

# # Prize Details
# Phasellus rutrum euismod turpis elementum ornare. Donec ut risus id ante gravida molestie. Integer cursus tempus porta. Sed vitae nunc quis nibh dapibus aliquet vel sed dolor. Donec id risus ut ipsum fermentum cursus quis sed massa. Nulla sit amet blandit orci, dapibus condimentum augue. Fusce suscipit purus et ultricies fermentum.

# # Requirements
# Mauris at est urna. Aenean ut elit venenatis augue dictum viverra:

# - **Nulla facilisi.** Donec vel justo odio. Vivamus consequat hendrerit arcu vel vestibulum. Proin malesuada mauris vitae nulla iaculis fringilla. 
# - **Proin tempor tempus ipsum id bibendum.** Duis vehicula nisi vel bibendum lacinia.
# - **Suspendisse libero dui**, hendrerit vitae eleifend sed, cursus ut tellus. Vivamus tristique, lectus in ullamcorper interdum, orci nisi vestibulum nisi, ac luctus est mi quis justo.
# - **Phasellus tempor laoreet felis a porta.** Aenean in sodales odio. Curabitur interdum bibendum orci, vitae hendrerit eros tempus at.